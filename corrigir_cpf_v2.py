#!/usr/bin/env python3
"""
Correção v2: CPF como texto no Excel
- Remove aspas do cadastro
- Força formato de texto ao salvar com openpyxl
"""

import pandas as pd
import shutil
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

BASE_DIR = Path("/home/thamiresvieira/projetos/Programa_mais_top")
BACKUP_DIR = BASE_DIR / "bases/Old/backup_cpf_texto_20250608"

def salvar_como_texto(df, caminho, colunas_texto):
    """Salva DataFrame forçando colunas específicas como texto no Excel"""
    wb = Workbook()
    ws = wb.active

    # Escrever header
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Escrever dados
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            col_name = df.columns[col_idx - 1]

            if col_name in colunas_texto and pd.notna(value):
                # Forçar como texto: prefixar com apóstrofo ou usar number_format
                cell.value = str(value)
                cell.number_format = '@'
            else:
                cell.value = value

    wb.save(caminho)

def limpar_cpf(valor):
    """Remove aspas e formata CPF"""
    if pd.isna(valor):
        return None
    s = str(valor).strip()
    # Remover aspas simples no início e fim
    s = s.strip("'")
    # Remover pontos, traços, barras
    s = s.replace('.', '').replace('-', '').replace('/', '').replace(' ', '')
    # Remover parte decimal
    if '.' in s:
        s = s.split('.')[0]
    # Preencher com zeros
    return s.zfill(11)

def processar_arquivo(caminho, colunas_cpf, abas=None):
    """Processa um arquivo Excel"""
    print(f"\n📁 {caminho.name}")

    xl = pd.ExcelFile(caminho)
    abas_processar = abas if abas else xl.sheet_names

    # Ler todas as abas
    dfs = {}
    for aba in xl.sheet_names:
        df = pd.read_excel(caminho, sheet_name=aba, dtype=str)

        if aba in abas_processar:
            print(f"   Aba '{aba}': {len(df)} linhas")
            for col in colunas_cpf:
                if col in df.columns:
                    df[col] = df[col].apply(limpar_cpf)
                    amostra = df[col].dropna().head(3).tolist()
                    print(f"      {col}: {amostra}")
        dfs[aba] = df

    # Salvar com openpyxl forçando texto
    with pd.ExcelWriter(caminho, engine='openpyxl') as writer:
        for aba, df in dfs.items():
            df.to_excel(writer, sheet_name=aba, index=False)

            # Forçar formato de texto nas colunas de CPF
            ws = writer.sheets[aba]
            for col_idx, col_name in enumerate(df.columns, 1):
                if col_name in colunas_cpf:
                    for row in range(2, len(df) + 2):
                        cell = ws.cell(row=row, column=col_idx)
                        if cell.value is not None:
                            cell.number_format = '@'

    print(f"   ✅ Salvo")

print("="*60)
print("CORREÇÃO V2: CPF COMO TEXTO NO EXCEL")
print("="*60)

# 1. Corrigir CADASTRO (remover aspas)
print("\n🔧 Corrigindo CADASTRO (removendo aspas)...")
df_cad = pd.read_excel(BASE_DIR / "bases/Tabelas/cadastro.xlsx", dtype=str)
df_cad.columns = df_cad.columns.str.strip().str.lower()
df_cad['cpf/cnpj'] = df_cad['cpf/cnpj'].apply(limpar_cpf)
print(f"   Amostra: {df_cad['cpf/cnpj'].head(5).tolist()}")

with pd.ExcelWriter(BASE_DIR / "bases/Tabelas/cadastro.xlsx", engine='openpyxl') as writer:
    df_cad.to_excel(writer, sheet_name='cadastro_2026-06-03T12_14_16.01', index=False)
    ws = writer.sheets['cadastro_2026-06-03T12_14_16.01']
    col_idx = list(df_cad.columns).index('cpf/cnpj') + 1
    for row in range(2, len(df_cad) + 2):
        ws.cell(row=row, column=col_idx).number_format = '@'

print("   ✅ Cadastro corrigido")

# 2. Processar outras bases
processar_arquivo(
    BASE_DIR / "bases/Tabelas/acessos.xlsx",
    colunas_cpf=["Cpf/Cnpj"]
)

processar_arquivo(
    BASE_DIR / "bases/Tabelas/pontos_por_cpf_.xlsx",
    colunas_cpf=["CPF", "CPFCNPJ", "DOCUMENTO"],
    abas=["CONSOLIDADO_RESGATES", "PONTOS POR CPF", "VALES", "LOJA VIRTUAL", "PAGAMENTO DE CONTAS", "RECARGA DE CELULAR"]
)

processar_arquivo(
    BASE_DIR / "bases/Tabelas/resgates_detalhado.xlsx",
    colunas_cpf=["CPF"]
)

processar_arquivo(
    BASE_DIR / "bases/Tabelas/Base_treinamentos.xlsx",
    colunas_cpf=["CPF", "Cnpj Distribuidor", "Cnpj PDV"]
)

# Vendas
print(f"\n📁 Vendas Processadas (todos os arquivos)")
vendas_dir = BASE_DIR / "bases/Vendas Processadas"
for arquivo in sorted(vendas_dir.glob("*.xlsx")):
    if arquivo.name.startswith("OLD") or arquivo.name.startswith("~"):
        continue
    try:
        processar_arquivo(arquivo, colunas_cpf=["CPF"])
    except Exception as e:
        print(f"   ❌ Erro em {arquivo.name}: {e}")

processar_arquivo(
    BASE_DIR / "bases/Tabelas/enquete.xlsx",
    colunas_cpf=["cpf"]
)

processar_arquivo(
    BASE_DIR / "bases/Tabelas/WHP_Aceite_Mensal_OUT_NOV_DEZ_2025_JAN_FEV_2026.xlsx",
    colunas_cpf=["CPF"]
)

processar_arquivo(
    BASE_DIR / "bases/Tabelas/WHP_PROD_Cadastro_Revenda_x_Loja_x_Regional.xlsx",
    colunas_cpf=["Cpf", "CNPJ_Revenda", "CNPJ_Loja"]
)

print("\n" + "="*60)
print("✅ CORREÇÃO CONCLUÍDA")
print("="*60)
